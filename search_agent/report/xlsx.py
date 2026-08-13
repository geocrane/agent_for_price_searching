# -*- coding: utf-8 -*-
"""Формирование итогового Excel-отчёта: строки из сводов → запись .xlsx (openpyxl).

Книга из ДВУХ листов:
  • «Итоги» — строка на позицию: Товар, Цена, Диапазон, уверенность, Комментарий по
    источникам. Единица измерения НЕ выводится.
  • «Детали» — строка на КАЖДУЮ найденную цену (и по одной на сайт, который цены не дал): запрос,
    домен, ссылка, статус, что нашлось именно там, и весь разбор отбора — учтена ли цена в итоге,
    на каком шаге отсеяна и почему, какой балл набрала и из чего он сложился.

Разбор отбора жил отдельным листом «Отбор», но замер показал, что его строки соответствуют
ценовым строкам «Деталей» один к одному (0 расхождений, 0 коллизий на реальном прогоне), а шесть
из одиннадцати его колонок дублировали соседний лист. Отдельный лист оправдан, когда у него своя
единица строки; здесь единица одна и та же — найденная цена. Два листа лишь заставляли сверять
глазами то, что помещается в одну строку.

Ещё раньше «Детали» показывали ПО ОДНОЙ цене на сайт, выбранной другим правилом
(`session.best_price`: точное, затем минимум), чем итог (взвешенный балл). Два правила в одной
книге расходились: у сайта с несколькими ценами в «Итогах» стояло одно число, в «Деталях» —
другое, и проверить итог по книге было нельзя. Теперь правило одно, а решение по каждой цене
берётся из `verdict.scored` — того же списка, по которому adjudicate выбирал.

Сборка строк — детерминированная (без модели). Комментарий подставляется отдельно
(report.rationale кладёт его в item['rationale']); если его нет — ячейка остаётся пустой.

Модуль shape-agnostic: свод (item['verdict']) принимается в ДВУХ формах —
  • полный дамп Verdict (primary — объект, excluded — список) — путь CLI/pipeline;
  • плоское событие `verdict` для UI (primary — число, excluded — счётчик) — путь webui.
`norm_verdict` приводит обе к единому виду.
"""
from ..obs.log import get_logger

log = get_logger("report")

# Заголовок «Комментарий модели» врал: при недоступной модели комментарий пишет код
# (report/rationale.py → fallback_rationale), и на прогоне 09.08 таких было 42%. Поэтому
# нейтральное название плюс отдельная колонка «кем составлен».
# Колонки «Номер» здесь нет намеренно: артикул больше не извлекается из входного файла
# (см. input/excel_reader.py), и колонка была бы пустой во всех строках.
COLUMNS = ["Товар", "Цена", "Тип", "Найденное название",
           "Минимальная цена", "Максимальная цена", "Источник цены", "Ссылка на источник",
           "Уверенность", "Подтверждено др. сайтами (±5%)", "Комментарий по источникам",
           "Комментарий составлен"]

# Пояснения к заголовкам: без них «Подтверждено сайтами = 0» читается как «цена плохая», хотя
# чаще это «розничный разброс больше 5%» или «после отсева осталась одна цена».
COLUMN_NOTES = {
    "Тип": "Точное/аналог называет МОДЕЛЬ, сверив товар со страницей.\n"
           "«НЕ ПРОВЕРЕНО МОДЕЛЬЮ» — цену взяли из структурной разметки страницы, потому что\n"
           "модель была недоступна: соответствие товару не сверял никто.",
    "Уверенность": "Итоговый балл цены, 0–100: тир источника, уверенность извлечения,\n"
                   "подтверждения, тип совпадения, наличие, близость к медиане, сайт вендора.",
    "Подтверждено др. сайтами (±5%)": (
        "Сколько ДРУГИХ доменов назвали цену в пределах ±5% от итоговой.\n"
        "Считается по ценам, прошедшим отбор (валюта → единица → аналог → выброс),\n"
        "а не по всем найденным предложениям.\n"
        "0 НЕ значит, что цена плохая: чаще это розничный разброс больше 5%\n"
        "или единственная цена, оставшаяся после отбора."),
    "Комментарий составлен": "«модель» — текст написала модель, сверив источники.\n"
                             "«сводка» — модель была недоступна, текст собран кодом из данных отбора.",
}
# Индексы колонок с числовой ценой (1-based) — записываем как числа, а не текст.
# Считаются от COLUMNS: Цена / Минимальная / Максимальная.
_PRICE_COLS = (2, 5, 6)
_CUR = {"RUB": "₽", "USD": "$", "EUR": "€", "CNY": "¥"}


# Лист «Детали»: строка на КАЖДЫЙ сайт выдачи. Здесь видна работа по товару целиком — что искали,
# с каким аффиксом, какой строкой это ушло в поиск, что нашлось на каждом сайте и почему. Итог на
# первом листе — следствие этих строк, и он должен быть проверяем без обращения к логам.
#
# «Название» — как в исходном файле, «Поисковый запрос» — то, что РЕАЛЬНО ушло в поиск. Две
# отдельные колонки намеренно: рядом сразу видно, не потерялось ли по дороге слово из наименования
# (в замере так обнаружилась пропажа бренда: «DELTA Аккумулятор Delta DTM 1212» → «Аккумулятор
# DTM 1212»). Одна колонка «как искали» этот разрыв прятала бы.
DETAIL_COLUMNS = ["Название", "Поисковый запрос", "Сайт", "Ссылка", "Статус",
                  "Найдено на сайте", "Тип", "Ед.", "Цена",
                  "Итог отбора", "Балл", "Из чего балл", "Комментарий"]
_DETAIL_PRICE_COLS = (9,)
_DETAIL_LINK_COL = 4
_DETAIL_WRAP_COLS = (1, 2, 4, 6, 10, 12, 13)
_DETAIL_WIDTHS = [38, 40, 20, 36, 16, 30, 9, 7, 11, 40, 8, 56, 44]

# Лист «Сайты»: снимок работы инструмента — строка на источник, с ОДНОЙ представительной ценой,
# как её показывал экран во время поиска. Разбор итога живёт на «Деталях», здесь его нет намеренно.
SITE_COLUMNS = ["Название", "Поисковый запрос", "Сайт", "Ссылка", "Статус",
                "Найдено на сайте", "Тип", "Цена (одна на сайт)", "Комментарий"]
_SITE_PRICE_COLS = (8,)
_SITE_LINK_COL = 4
_SITE_WRAP_COLS = (1, 2, 4, 6, 9)
_SITE_WIDTHS = [38, 46, 22, 40, 16, 32, 11, 15, 54]


def _match_str(m) -> str | None:
    """Значение MatchType из enum/строки/None."""
    if m is None:
        return None
    return getattr(m, "value", None) or str(m)


def money(value, currency: str = "RUB") -> str:
    """Число + символ валюты, целые без дробной части («430 ₽», «1 250.50 ₽»)."""
    if value is None:
        return ""
    sym = _CUR.get((currency or "RUB").upper())
    try:
        f = float(value)
    except (TypeError, ValueError):
        return str(value)
    body = ("%d" % f) if f.is_integer() else ("%.2f" % f)
    return "%s %s" % (body, sym) if sym else "%s %s" % (body, currency)


def norm_verdict(v: dict | None) -> dict:
    """Привести свод (полный Verdict-дамп ИЛИ плоское событие) к единому виду для отчёта."""
    if not v:
        return {"value": None, "currency": "RUB", "match": None, "domain": None,
                "url": None, "found": None, "unit": "", "market_corridor": False,
                "price_min": None, "price_max": None, "confidence": None,
                "corroborated_by": 0, "excluded_list": [], "excluded_count": 0,
                "scored": [], "tie": False, "needs_review": False, "verified": True}
    primary = v.get("primary")
    if isinstance(primary, dict):                    # полный дамп Verdict
        value = primary.get("value")
        match = _match_str(primary.get("match"))
        domain = primary.get("source_domain")
        url = primary.get("url")
        found = primary.get("found")
        verified = primary.get("verified", True)
    else:                                            # плоское событие verdict
        value = primary
        match = _match_str(v.get("match"))
        domain = v.get("domain")
        url = v.get("url")
        found = v.get("found")
        verified = v.get("verified", True)
    excluded = v.get("excluded")
    excluded_list = excluded if isinstance(excluded, list) else []
    excluded_count = len(excluded_list) if isinstance(excluded, list) else int(excluded or 0)
    return {
        "value": value,
        "currency": (v.get("currency") or "RUB"),
        "match": match,
        "domain": domain,
        "url": url,
        "found": found,
        "price_min": v.get("price_min"),
        "price_max": v.get("price_max"),
        "unit": (primary.get("unit") if isinstance(primary, dict) else None) or v.get("unit") or "",
        "market_corridor": bool(v.get("market_corridor")),
        "confidence": v.get("confidence"),
        "corroborated_by": int(v.get("corroborated_by") or 0),
        "excluded_list": excluded_list,
        "excluded_count": excluded_count,
        # Решение по каждой цене — источник листа «Отбор». У плоского события его нет: webui
        # шлёт свод без разбора, и лист тогда просто останется без строк этой позиции.
        "scored": v.get("scored") or [],
        "tie": bool(v.get("tie")),
        "needs_review": bool(v.get("needs_review")),
        # Цену назвала модель, сверив товар со страницей? False — взята из структурной разметки
        # при недоступной модели. Отчёт обязан это показывать: «точным» такое называть нельзя.
        "verified": bool(verified) if verified is not None else True,
    }


def to_int(value) -> int | None:
    """Целое число цены (int) или None. Отчёт хранит цены как ЧИСЛА (сортируемо/суммируемо)."""
    if value is None:
        return None
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return None


def match_cell(nv: dict, not_found_reason: str | None, on_request: dict | None = None) -> str:
    """Ячейка «Тип»: что за цена и за что она.

    Кроме точное/аналог здесь же называется единица («за м3») и честно помечается коридор рынка:
    у родовой позиции крайние значения — это разброс КАТЕГОРИИ, поэтому в колонках мин/макс стоят
    границы срединного коридора, и пользователь должен видеть, что это именно они.
    """
    if nv["value"] is None:
        if on_request:
            return "цена по запросу — %s" % (on_request.get("match") or "точное")
        return "не найдено — %s" % (not_found_reason or "нет данных")
    # «точное» пишем ТОЛЬКО когда товар со страницей сверила модель. Цена из структурной
    # разметки (модель была недоступна) раньше печаталась как «точное» — читатель не мог
    # отличить проверенную цену от непроверенной, и 32% отчёта 09.08 были именно такими.
    parts = ["НЕ ПРОВЕРЕНО МОДЕЛЬЮ (цена из разметки страницы)"] if not nv.get("verified", True) \
        else [nv["match"] or "точное"]
    if nv.get("unit"):
        parts.append("за %s" % nv["unit"])
    if nv.get("market_corridor"):
        parts.append("мин/макс — коридор рынка 25–75%")
    # Аналог, которого никто не подтвердил, найденный на единственном сайте, — это цена ПОХОЖЕГО
    # товара другого бренда. Число в отчёте остаётся, но брать его в смету без проверки нельзя.
    if nv.get("needs_review") and nv.get("verified", True):
        parts.append("ТРЕБУЕТ ПРОВЕРКИ — аналог с одного сайта, никем не подтверждён")
    if nv.get("tie"):
        parts.append("ничья по баллам — см. лист «Отбор»")
    return " · ".join(parts)


def confidence_pct(nv: dict) -> int | None:
    """Уверенность в итоге, 0–100. Доля от единицы читается хуже, чем привычные проценты."""
    c = nv.get("confidence")
    if c is None:
        return None
    try:
        return int(round(float(c) * 100))
    except (TypeError, ValueError):
        return None


def build_row(item: dict) -> dict:
    """Собрать одну строку отчёта из товара (со сводом). Детерминированно, дискретные колонки."""
    nv = norm_verdict(item.get("verdict"))
    nfr = item.get("not_found_reason")
    onreq = item.get("on_request") or None
    price = to_int(nv["value"])
    # «Цена по запросу» — это НАЙДЕННЫЙ товар без цифры: показываем, что именно нашли и где.
    return {
        "row": item.get("row"),
        "Товар": item.get("name") or "",
        "Цена": price,                                          # число (int) или None
        "Тип": match_cell(nv, nfr, onreq),
        "Найденное название": (nv["found"] or (onreq or {}).get("found") or "").strip(),
        "Минимальная цена": to_int(nv["price_min"]),
        "Максимальная цена": to_int(nv["price_max"]),
        "Источник цены": nv["domain"] or (onreq or {}).get("domain") or "",
        "Ссылка на источник": nv["url"] or (onreq or {}).get("url") or "",
        "Уверенность": confidence_pct(nv),
        "Подтверждено др. сайтами (±5%)": nv["corroborated_by"] if price is not None else None,
        "Комментарий по источникам": (item.get("rationale") or "").strip(),
        "Комментарий составлен": item.get("rationale_by") or "",
        # Подсветка строки: цены нет ЛИБО она есть, но требует проверки (одинокий аналог).
        "_found": ((price is not None or bool(onreq))
                   and not nv.get("needs_review") and nv.get("verified", True)),
    }


def build_rows(items: list[dict]) -> list[dict]:
    return [build_row(it) for it in items]


def item_query(item: dict) -> str:
    """Реальная поисковая строка позиции. `queries` — запас для сессий прежних версий."""
    q = item.get("query")
    if q:
        return str(q)
    queries = item.get("queries") or []
    if isinstance(queries, (list, tuple)):
        return str(queries[0]) if queries else ""
    return str(queries)


def _decisions_index(nv: dict) -> dict:
    """Решения свода по ключу (url, цена) — чтобы к каждой цене приписать её судьбу в итоге."""
    idx = {}
    for d in nv.get("scored") or []:
        try:
            idx[(d.get("url") or "", round(float(d.get("value")), 2))] = d
        except (TypeError, ValueError):
            continue
    return idx


def pick_cell(d: dict | None, *, legacy: bool = False) -> str:
    """Колонка «Итог отбора»: судьба конкретной цены словами, вместе с причиной отсева.

    Раньше это были две колонки на двух листах («Учтено в итоге» и «Статус и причина»), но обе
    описывают одну и ту же строку — найденную цену, — и разносить их было незачем.
    """
    if d is None:
        # Свод прежней версии разбора не содержит. Пустая ячейка выглядела бы как сбой, поэтому
        # причина называется прямо (правило «ничего не произошло недопустимо»).
        return "разбор недоступен — свод прежней версии, повторите поиск" if legacy else ""
    if d.get("is_primary"):
        return "★ ИТОГ"
    if d.get("accepted"):
        return "учтена в сравнении" + (", %s" % d["reason"] if d.get("reason") else "")
    return "отсеяна на шаге %s (%s): %s" % (
        d.get("step"), d.get("step_name") or "", d.get("reason") or "")


def build_detail_rows(items: list[dict]) -> list[dict]:
    """Строки листа «Детали»: по одной на КАЖДУЮ найденную цену, с полным разбором отбора.

    Раньше здесь была строка на сайт, а цена бралась через `session.best_price` («точное, затем
    минимум»). У сайта с несколькими ценами (карточка с блоком «похожие товары», листинг
    маркетплейса) в лист попадала одна из них — и почти никогда та, что стояла в «Итогах».
    Итог оказывался непроверяемым по книге.

    Сайты, цен не давшие (заблокированные, пустые, пропущенные, «цена по запросу»), дают по одной
    строке НАРАВНЕ с удачными: пользователь должен видеть, что было проверено, а не только то, что
    получилось (правило «логировать всё и видимо»). Для них берётся сводка
    `session.summarize_cand` — та же функция, которой рисуется таблица выдачи на экране.

    Название и запрос дублируются в каждой строке позиции — так лист читается и фильтруется сам
    по себе, без сверки с первым листом.
    """
    from ..extract.adjudicate import parts_formula
    from ..session import status_label, summarize_cand
    out: list[dict] = []
    for it in items:
        head = {"Название": it.get("name") or "", "Поисковый запрос": item_query(it)}
        nv = norm_verdict(it.get("verdict"))
        idx = _decisions_index(nv)
        # Свод прежней версии разбора не содержит — ячейки объясняют это вместо того, чтобы
        # молча пустовать.
        legacy = bool(it.get("verdict")) and not nv.get("scored")
        for c in it.get("candidates", []) or []:
            v = summarize_cand(c)
            base = {**head, "Сайт": c.get("domain") or "", "Ссылка": c.get("url") or "",
                    "Статус": status_label(v["status"])}
            prices = c.get("prices") or []
            if not prices:
                out.append({
                    **base,
                    "Найдено на сайте": (v["found"] or "").strip(),
                    "Тип": v["match"] or ("по запросу" if v["on_request"] else ""),
                    "Ед.": "",
                    "Цена": to_int(v["price"]),
                    "Итог отбора": "",                # цены нет — отбирать нечего
                    "Балл": None, "Из чего балл": "",
                    # Комментарий модели по источнику; нет — видимая причина (почему пусто).
                    "Комментарий": (v["comment"] or v["note"] or "").strip(),
                })
                continue
            for p in prices:
                value = p.get("value")
                key = None
                try:
                    key = (p.get("url") or c.get("url") or "", round(float(value), 2))
                except (TypeError, ValueError):
                    pass
                d = idx.get(key)
                score = (d or {}).get("score")
                out.append({
                    **base,
                    "Найдено на сайте": (p.get("found") or "").strip(),
                    "Тип": _match_str(p.get("match")) or "",
                    "Ед.": p.get("unit") or "",
                    "Цена": to_int(value),
                    "Итог отбора": pick_cell(d, legacy=legacy),
                    "Балл": None if score is None else round(float(score) * 100, 1),
                    "Из чего балл": parts_formula((d or {}).get("parts") or {}),
                    "Комментарий": (p.get("snippet") or v["note"] or "").strip(),
                })
    return out


def build_site_rows(items: list[dict]) -> list[dict]:
    """Строки листа «Сайты»: по одной на источник — ровно то, что было видно на экране.

    Это снимок работы инструмента, а не разбор итога: у сайта показана ОДНА представительная цена
    по правилу экрана (`session.best_price`: точное совпадение, затем минимальная). Правило это
    НЕ совпадает с выбором итога, и когда такой лист назывался «Детали», расхождение читалось как
    ошибка отчёта. Здесь у листа честная роль — «что показывалось в ходе поиска», а полный список
    цен с разбором отбора лежит на «Деталях».
    """
    from ..session import status_label, summarize_cand
    out: list[dict] = []
    for it in items:
        head = {"Название": it.get("name") or "", "Поисковый запрос": item_query(it)}
        for c in it.get("candidates", []) or []:
            v = summarize_cand(c)
            out.append({
                **head,
                "Сайт": c.get("domain") or "",
                "Ссылка": c.get("url") or "",
                "Статус": status_label(v["status"]),
                "Найдено на сайте": (v["found"] or "").strip(),
                "Тип": v["match"] or ("по запросу" if v["on_request"] else ""),
                "Цена (одна на сайт)": to_int(v["price"]),
                "Комментарий": (v["comment"] or v["note"] or "").strip(),
            })
    return out


def _write_sheet(ws, columns: list[str], rows: list[dict], *, widths: list[int],
                 wrap_cols: tuple, price_cols: tuple, link_col: int | None,
                 price_col: int | None) -> None:
    """Наполнить лист: шапка, границы, переносы, числовые цены, ссылки, подсветка «без цены».

    Оформление у обоих листов одно и то же и живёт ЗДЕСЬ: два экземпляра стилевого кода
    неминуемо разъедутся, и книга станет выглядеть как склейка из двух разных отчётов.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    notfound_fill = PatternFill("solid", fgColor="FCE4E4")     # мягкий красный — цена не найдена
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")
    link_font = Font(color="1F4E78", underline="single")

    from openpyxl.comments import Comment

    for col, name in enumerate(columns, start=1):              # строка 1 — заголовки
        c = ws.cell(row=1, column=col, value=name)
        c.fill, c.font, c.border = header_fill, header_font, border
        c.alignment = Alignment(vertical="center", wrap_text=True)
        # Примечание к заголовку: формула колонки должна читаться из самого отчёта, а не
        # угадываться. «Подтверждено сайтами = 0» без пояснения читается как «цена плохая».
        note = COLUMN_NOTES.get(name)
        if note:
            c.comment = Comment(note, "search_agent", height=150, width=380)
    ws.freeze_panes = "A2"

    for i, r in enumerate(rows, start=2):
        for col, name in enumerate(columns, start=1):
            val = r.get(name)
            c = ws.cell(row=i, column=col, value=val)
            c.border = border
            c.alignment = wrap_top if col in wrap_cols else top
            if col in price_cols and isinstance(val, int):
                c.number_format = "0"                # чистое целое (без валюты — сортируемо/суммируемо)
            if col == link_col and val:              # ссылка → кликабельная гиперссылка
                c.hyperlink = val
                c.font = link_font
        if price_col and not r.get("_found"):        # подсветить строку без надёжной цены
            ws.cell(row=i, column=price_col).fill = notfound_fill

    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def write_xlsx(rows: list[dict], path: str, *, sheet_title: str = "Итоги",
               detail_rows: list[dict] | None = None, detail_title: str = "Детали",
               site_rows: list[dict] | None = None, site_title: str = "Сайты") -> str:
    """Записать отчёт в .xlsx: итоги по позициям + все цены с разбором отбора + снимок по сайтам.

    Возвращает путь к файлу. Требует openpyxl (уже в зависимостях, см. input.excel_reader).
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    _write_sheet(ws, COLUMNS, rows,
                 widths=[40, 11, 34, 34, 13, 13, 20, 40, 12, 15, 60, 22],
                 wrap_cols=(1, 3, 4, 11),            # Товар / Тип / Найденное название / Комментарий
                 price_cols=_PRICE_COLS, link_col=8, price_col=2)

    if detail_rows is not None:
        ws2 = wb.create_sheet(detail_title)
        _write_sheet(ws2, DETAIL_COLUMNS, detail_rows,
                     widths=_DETAIL_WIDTHS, wrap_cols=_DETAIL_WRAP_COLS,
                     price_cols=_DETAIL_PRICE_COLS, link_col=_DETAIL_LINK_COL, price_col=None)
        ws2.auto_filter.ref = ws2.dimensions          # по листу деталей фильтруют и сортируют

    if site_rows is not None:
        ws3 = wb.create_sheet(site_title)
        _write_sheet(ws3, SITE_COLUMNS, site_rows,
                     widths=_SITE_WIDTHS, wrap_cols=_SITE_WRAP_COLS,
                     price_cols=_SITE_PRICE_COLS, link_col=_SITE_LINK_COL, price_col=None)
        ws3.auto_filter.ref = ws3.dimensions

    wb.save(path)
    found = sum(1 for r in rows if r["_found"])
    log.info("Отчёт записан: %s (позиций=%d, с ценой=%d, строк цен=%d, сайтов=%d)",
             path, len(rows), found, len(detail_rows or []), len(site_rows or []))
    return path
